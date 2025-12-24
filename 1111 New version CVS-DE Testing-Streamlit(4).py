# -*- coding: utf-8 -*-
"""
1111 New CVS-DE Testing-Streamlit.py
Streamlit app for Classical Variable Sampling (CVS-DE)
- Bilingual UI (中文/English)
- Upload Excel with sheet 'CVS-DE'
- Modes: Symmetric / Asymmetric / Both
- CPI uses sample SD (ddof=1), symmetric halfwidth = CPI (Excel-consistent)
- Export fixed path: r"H:\VS Code-Auditing\output-CVS-DE.xlsx" + browser download
- Teaching Mode (bottom, collapsible): Variable Glossary + numbered formulas (Symmetric only)
Run:
    streamlit run "1111 New CVS-DE Testing-Streamlit.py"
"""

import math
import os
# from turtle import color
from typing import Any, Dict, Optional, Tuple
import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import webbrowser
import qrcode
from PIL import Image, ImageDraw, ImageFont
import io

import streamlit as st

st.write("STREAMLIT BOOT OK")
st.stop()

# ============================
# 產生帶 Logo 的 QR Code
# ============================
def make_qr(url: str, logo_path: str = None, size: int = 600):
    qr = qrcode.QRCode(
        version=2,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=12,
        border=2,
    )
    qr.add_data(url)
    qr.make(fit=True)

    img_qr = qr.make_image(fill_color="black", back_color="white").convert("RGB")

    # 如果沒有 logo，直接回傳 QR
    if not logo_path:
        return img_qr

    # 插入 Logo
    logo = Image.open(logo_path)
    w, h = img_qr.size
    factor = 5
    logo = logo.resize((w // factor, h // factor))

    pos = ((w - logo.size[0]) // 2, (h - logo.size[1]) // 2)
    img_qr.paste(logo, pos, mask=logo if logo.mode == "RGBA" else None)
    return img_qr


# ============================
# Streamlit UI
# ============================
st.set_page_config(page_title="Green Audit – QR Code", layout="centered")

st.markdown(
    """
    <div style="text-align:center;">
        <h1 style="color:#2c7a7b;">🌿 Green Audit — Access Page</h1>
        <p style="font-size:18px; color:#444;">
            Scan the QR Code below to open the interactive auditing system.<br>
            The system supports mobile browsing and real-time computations.
        </p>
    </div>
    """,
    unsafe_allow_html=True
)

# 你的部署網址放這裡
URL = "https://2025-cross-disciplinary-creative-programming-competition-3fzxs.streamlit.app/"

qr_img = make_qr(URL)

# 將 QR Code 輸出到 Streamlit
buf = io.BytesIO()
qr_img.save(buf, format="PNG")
st.image(buf.getvalue(), width=350)

# ============================
# 美觀卡片式外框
# ============================
st.markdown(
    """
    <div style="
        margin-top:30px;
        padding:20px;
        border-radius:12px;
        background:linear-gradient(135deg, #e6fffa, #f0fff4);
        border: 2px solid #38a169;
        text-align:center;
    ">
        <h3 style="color:#2f855a;">📲 Mobile Friendly • 🚀 Cloud Deployed • 🔐 Secure Access</h3>
        <p style="font-size:16px; color:#333;">
            Use your smartphone to access the auditing platform instantly.
            <br>
            QR Code includes error correction (H-Level) for reliable scanning.
        </p>
    </div>
    """,
    unsafe_allow_html=True
)

# ==============================
# 其餘主程式邏輯、互動、顯示
# ==============================

# ===== Fixed export path (per your requirement) =====
FIXED_OUT_PATH = r"J:\VS Code-Auditing\output-CVS-DE.xlsx"

# ===== CL → (ZA, ZR) mapping (ZA: ARIA column, ZR: ARIR column) =====
CONF_LEVEL_MAP: Dict[str, Tuple[float, float]] = {
    "99": (2.33, 2.58),
    "95": (1.64, 1.96),
    "90": (1.28, 1.64),
    "80": (0.84, 1.28),
    "75": (0.67, 1.15),
}

# ===== Formatting helpers =====
def fmt2(x) -> str:
    try:
        return f"{float(x):,.2f}"
    except Exception:
        return str(x)

def fmt4(x) -> str:
    try:
        return f"{float(x):,.4f}"
    except Exception:
        return str(x)

# ===== Load sample block from uploaded Excel (sheet 'CVS-DE') =====
def load_sample_block_from_uploaded(file_buffer):
    raw = pd.read_excel(file_buffer, sheet_name="CVS-DE", header=None)

    def row_has_header(sr: pd.Series) -> bool:
        s = sr.astype(str).str.lower()
        return s.str.contains("Account number", na=False).any() or s.str.contains("account number", na=False).any()

    header_idx_list = raw.index[raw.apply(row_has_header, axis=1)].tolist()
    if not header_idx_list:
        raise RuntimeError("找不到表頭列（需含 'Account NO.' 或 'Account Number'）")
    header_idx = header_idx_list[0]

    data = raw.iloc[header_idx + 1:, 0:4].copy()
    data.columns = [
        "Account Number",
        "Recorded Accounts",
        "Audited Accounts",
        "Factual Misstatement (ej)",
    ]

    stop_idx = data.index[data["Account Number"].astype(str).str.contains("Total", case=False, na=False)]
    if len(stop_idx) > 0:
        data = data.loc[: stop_idx[0] - 1]

    # Cast numeric
    for c in ["Recorded Accounts", "Audited Accounts", "Factual Misstatement (ej)"]:
        data[c] = pd.to_numeric(data[c], errors="coerce")

    e_series = data["Factual Misstatement (ej)"].dropna()
    m = int(e_series.shape[0])
    sum_e = float(e_series.sum())
    visible_rows = int(data.shape[0])
    return data, e_series, m, sum_e, visible_rows

# ===== Core compute (Excel-consistent CPI & intervals) =====
def compute_all(
    SD_star: float, ZA: float, ZR: float, N: int, TM: float, E_star: float,
    e_series: pd.Series, m: int, sum_e: float, n_override: Optional[int]
) -> Dict[str, Any]:

    if TM <= E_star:
        raise ValueError("TM must be greater than E* to compute sample size (avoid division by zero).")

    # Auto sample size by formula
    n_auto_exact = ((SD_star * (ZA + ZR) * N) / (TM - E_star)) ** 2
    n_auto_int = math.ceil(n_auto_exact)
    n = int(n_override) if (isinstance(n_override, int) and n_override >= 1) else n_auto_int
    if m > n:
        raise ValueError(f"Count of nonzero errors m={m} cannot exceed chosen sample size n={n}.")

    # Point estimate: ē = Σe / n (fill zero for non-error draws)
    e_bar = sum_e / n
    e_full = np.concatenate([e_series.values, np.zeros(max(0, n - m))])
    SD_sample = float(np.std(e_full, ddof=1)) if n > 1 else 0.0  # Excel STDEV.S

    # SE(mean) using SD* with FPC
    SE_mean = SD_star / math.sqrt(n) * math.sqrt((N - n) / N)

    # Total misstatement
    T_hat = N * e_bar

    # CPI (amount) — Excel-consistent: use sample SD (not SD*)
    CPI_amt = N * ZA * SD_sample / math.sqrt(n) * math.sqrt((N - n) / N)

    # Symmetric interval — use CPI as halfwidth
    Z_sym = (ZA + ZR) / 2.0  # display only
    half_sym = CPI_amt
    L_sym = T_hat - half_sym
    U_sym = T_hat + half_sym
    dec_sym = "Accept" if U_sym <= TM else "Reject"

    # Asymmetric interval — halfwidths with ZA/ZR * (N*SE_mean)
    half_L = ZA * (N * SE_mean)
    half_R = ZR * (N * SE_mean)
    L_asym = T_hat - half_L
    U_asym = T_hat + half_R
    dec_asym = "Accept" if U_asym <= TM else "Reject"

    return {
        "n_auto_exact": n_auto_exact,
        "n_auto_int": n_auto_int,
        "n_used": n,
        "inputs": {"SD_star": SD_star, "ZA": ZA, "ZR": ZR, "N": N, "TM": TM, "E_star": E_star, "m": m, "sum_e": sum_e},
        "e_bar": e_bar,
        "SD_sample": SD_sample,
        "SE_mean": SE_mean,
        "T_hat": T_hat,
        "CPI_amt": CPI_amt,
        "sym": {"Z_sym": Z_sym, "L": L_sym, "U": U_sym, "half": half_sym, "decision": dec_sym},
        "asym": {"ZA": ZA, "ZR": ZR, "L": L_asym, "U": U_asym, "half_L": half_L, "half_R": half_R, "decision": dec_asym},
    }

# ===== Results table (bilingual) =====
def build_results_df(res: Dict[str, Any]) -> pd.DataFrame:
    sym, asym = res["sym"], res["asym"]
    rows = [
        ["樣本量", "Sample Size (n)", f"{res['n_used']:,}"],
        ["預設樣本量（公式）", "Default n by formula", fmt4(res["n_auto_exact"]) + f" → ceil={res['n_auto_int']}"],
        ["母體大小", "Population Size (N)", f"{res['inputs']['N']:,}"],
        ["有錯樣本數", "Count of Nonzero Errors (m)", f"{res['inputs']['m']:,}"],
        ["誤差總和", "Sum of Errors (Σe_j)", fmt4(res['inputs']['sum_e'])],
        ["樣本標準差", "Sample SD", fmt4(res["SD_sample"])],
        ["母體平均點估計(誤差)", "Point estimate of population mean (μ̂=ē)", fmt4(res["e_bar"])],
        ["總誤差點估計", "Total misstatement (T̂=N*ē)", fmt2(res["T_hat"])],
        ["平均之標準誤", "SE of mean (with FPC)", fmt4(res["SE_mean"])],
        ["精確度區間（金額）", "Precision interval (CPI, amount)", fmt2(res["CPI_amt"])],
        ["對稱 Z 值（顯示）", "Z_sym = (ZA+ZR)/2 (display)", fmt4(sym["Z_sym"])],
        ["對稱區間下限", "Symmetric lower", fmt2(sym["L"])],
        ["對稱區間上限", "Symmetric upper", fmt2(sym["U"])],
        ["對稱半寬（=CPI）", "Symmetric halfwidth (=CPI)", fmt2(sym["half"])],
        ["對稱決策", "Symmetric decision", sym["decision"]],
        ["非對稱 ZA（下限）", "Asymmetric ZA (lower)", fmt4(asym["ZA"])],
        ["非對稱 ZR（上限）", "Asymmetric ZR (upper)", fmt4(asym["ZR"])],
        ["非對稱區間下限", "Asymmetric lower", fmt2(asym["L"])],
        ["非對稱區間上限", "Asymmetric upper", fmt2(asym["U"])],
        ["非對稱半寬(左)", "Asymmetric halfwidth (left)", fmt2(asym["half_L"])],
        ["非對稱半寬(右)", "Asymmetric halfwidth (right)", fmt2(asym["half_R"])],
        ["非對稱決策", "Asymmetric decision", asym["decision"]],
    ]
    return pd.DataFrame(rows, columns=["中文欄位名稱", "English Label", "Value"])

# ===== Workpapers (add ej^2 + Total) =====
def build_workpapers_df(data: pd.DataFrame) -> pd.DataFrame:
    wp = data.copy()
    ej2 = (wp["Factual Misstatement (ej)"].astype(float)) ** 2
    wp["Factual Misstatement^2 (ej^2)"] = ej2
    total_row = {
        "Account Number": "Total",
        "Recorded Accounts": wp["Recorded Accounts"].sum(skipna=True),
        "Audited Accounts": wp["Audited Accounts"].sum(skipna=True),
        "Factual Misstatement (ej)": wp["Factual Misstatement (ej)"].sum(skipna=True),
        "Factual Misstatement^2 (ej^2)": wp["Factual Misstatement^2 (ej^2)"].sum(skipna=True),
    }
    wp = pd.concat([wp, pd.DataFrame([total_row])], ignore_index=True)
    return wp

# ===== Export to Excel (accounting format) =====
def export_to_excel(results_df: pd.DataFrame, work_df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "results"
    for r in dataframe_to_rows(results_df, index=False, header=True):
        ws1.append(r)

    ws2 = wb.create_sheet("workpapers")
    for r in dataframe_to_rows(work_df, index=False, header=True):
        ws2.append(r)

    # Apply accounting-like format to numeric-ish cells
    def set_thousand(ws):
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                try:
                    float(cell.value)
                    cell.number_format = '#,##0.00'
                except Exception:
                    pass

    set_thousand(ws1)
    set_thousand(ws2)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

# ===== Streamlit UI =====
st.set_page_config(page_title="CVS-DE Auditing App", layout="wide")
st.title("審計抽樣（傳統變量抽樣）Classical Variable Sampling — CVS-DE")

with st.sidebar:
    st.header("上傳資料 / Upload Excel")
    file = st.file_uploader("請上傳含 CVS-DE 工作表的 Excel 檔案 / Upload Excel with sheet 'CVS-DE'", type=["xlsx"])

    st.header("參數設定 / Parameters")
    SD_star = st.number_input("估計母體標準差 SD★ / Estimated population SD (SD*)", min_value=0.0, value=20.0, step=0.1)
    N = st.number_input("母體大小 N / Population size N", min_value=1, value=4000, step=1)
    TM = st.number_input("可容忍錯誤 TM / Tolerable misstatement (total)", min_value=0.0, value=21000.0, step=100.0)
    E_star = st.number_input("預期母體錯誤 E★ / Expected misstatement (total)", min_value=0.0, value=1500.0, step=100.0)

    st.markdown("---")
    st.subheader("信賴水準（分開設定）/ Confidence Levels")
    cl_aria = st.selectbox("ARIA Confidence Level（ZA）", ["", "75", "80", "90", "95", "99"], index=3, help="空白 = 手動輸入 ZA")
    cl_arir = st.selectbox("ARIR Confidence Level（ZR）", ["", "75", "80", "90", "95", "99"], index=2, help="空白 = 手動輸入 ZR")

    ZA_default = CONF_LEVEL_MAP[cl_aria][0] if cl_aria else 1.28
    ZR_default = CONF_LEVEL_MAP[cl_arir][1] if cl_arir else 1.64
    ZA = st.number_input("ZA（若 CL 空白則手動輸入）", min_value=0.0, value=ZA_default, step=0.01)
    ZR = st.number_input("ZR（若 CL 空白則手動輸入）", min_value=0.0, value=ZR_default, step=0.01)
    n_override_txt = st.text_input("確認樣本量 n（留白=使用自動估計）/ Confirm n (blank = use auto)")
    n_override = int(n_override_txt) if n_override_txt.strip().isdigit() else None

    # ==== Sample Size Auto Calculation (show before user input) ⭐ 自動估計樣本量（放在 Sidebar，Sample size 區塊前方） ====
    st.subheader("📌 自動估計樣本量 / Auto-calculated Sample Size")
    if SD_star > 0 and N > 0 and TM > E_star:
        try:
            n_auto_exact_pre = ((SD_star * (ZA + ZR) * N) / (TM - E_star)) ** 2
            n_auto_int_pre = math.ceil(n_auto_exact_pre)

            st.markdown(
                f"""
                **🔍 依公式自動估計 / Estimated by formula：**

                - **n（精確值 / exact） = `{n_auto_exact_pre:.4f}`**  
                - **n（無條件進位 / ceil） = `{n_auto_int_pre}`**

                （此為預設樣本量，下方輸入留白即採用此值）
                """
            )
        except Exception:
            st.warning("無法計算樣本量，請檢查參數輸入。")
    else:
        st.info("請先輸入 SD★、N、TM、E★、ZA、ZR，以自動計算樣本量。")

    # ---- 👉 樣本量輸入區 ----# ---- 🔎 查核人員最終決定抽樣計畫 ----
    st.markdown("### 🔎 查核人員最終決定抽樣計畫 / Auditor’s Final Sampling Plan")
    st.subheader("👉 樣本量 / Sample size")
    # n_override_txt = st.text_input("確認樣本量 n（留白=使用自動估計）/ Confirm n (blank = use auto)")
    st.markdown("---")
    mode = st.radio("顯示模式 / Display mode", options=["1 對稱 Symmetric", "2 非對稱 Asymmetric", "3 並列 Both"], index=2)

    run = st.button("執行分析 / Run Analysis")

# === UI 美化：Green Audit 主題 ===
st.markdown("""
<style>
    .main {
        background-color: #F9FFFB;
    }
    div[data-testid="stSidebar"] {
        background-color: #E7F7EE;
    }
    .metric-container {
        background: #ffffff;
        padding: 10px;
        border-radius: 10px;
        border: 1px solid #d8e8df;
        margin-bottom: 10px;
    }
</style>
""", unsafe_allow_html=True)

if run:
    if file is None:
        st.error("請先上傳 Excel。Please upload an Excel file first.")
        st.stop()

    # Load
    try:
        data, e_series, m, sum_e, rows = load_sample_block_from_uploaded(file)
    except Exception as e:
        st.error(f"讀取樣本區塊失敗：{e}")
        st.stop()

    # Compute
    try:
        res = compute_all(
            SD_star=SD_star, ZA=ZA, ZR=ZR, N=N, TM=TM, E_star=E_star,
            e_series=e_series, m=m, sum_e=sum_e, n_override=n_override
        )
    except Exception as e:
        st.error(f"計算失敗：{e}")
        st.stop()

    st.info(f"預設樣本量（公式）Default n by formula = {res['n_auto_exact']:.4f} → ceil = {res['n_auto_int']}")

    # Key metrics
    st.subheader("重點結果 / Key Results")
    colA, colB, colC, colD, colE = st.columns(5)
    colA.metric("樣本標準差 Sample SD", fmt4(res["SD_sample"]))
    colB.metric("ē（點估計）", fmt4(res["e_bar"]))
    colC.metric("T̂ = N*ē", fmt2(res["T_hat"]))
    colD.metric("SE(mean) with FPC", fmt4(res["SE_mean"]))
    colE.metric("CPI（金額）", fmt2(res["CPI_amt"]))

    # Intervals
    def decision_html(decision):
        color = "red" if "Reject" in decision else "green"
        return f"<span style='color:{color}; font-weight:700; font-size:1.3em'>{decision}</span>"

    if mode.startswith("1"):
        st.markdown("### 對稱區間 / Symmetric Interval")

        sym = res["sym"]

        st.markdown(
            f"""
        <div style="font-size:1.15em; line-height:1.6">
        下限 <b>Lower</b> = {fmt2(sym['L'])} <br>
        上限 <b>Upper</b> = {fmt2(sym['U'])} <br>
        半寬 <b>Halfwidth (CPI)</b> = {fmt2(sym['half'])} <br>
        最終決策 <b>Decision</b> = {decision_html(sym['decision'])}
        </div>
        """,
            unsafe_allow_html=True,
        )
    elif mode.startswith("2"):
        st.markdown("### 非對稱區間 / Asymmetric Interval")

        asym = res["asym"]

        st.markdown(
            f"""
        <div style="font-size:1.15em; line-height:1.6">
        下限 <b>Lower</b> = {fmt2(asym['L'])} <br>
        上限 <b>Upper</b> = {fmt2(asym['U'])} <br>
        左半寬 <b>Left</b> = {fmt2(asym['half_L'])} <br>
        右半寬 <b>Right</b> = {fmt2(asym['half_R'])} <br>
        最終決策 <b>Decision</b> = {decision_html(asym['decision'])}
        </div>
        """,
            unsafe_allow_html=True,
        )
    else:
        st.markdown("### 並列比較 / Both")

        sym, asym = res["sym"], res["asym"]

        st.markdown(
            f"""
        <div style="font-size:1.1em; line-height:1.6">
        <b>[對稱]</b><br>
        Lower = {fmt2(sym['L'])} , Upper = {fmt2(sym['U'])} , Half = {fmt2(sym['half'])} <br>
        Decision = {decision_html(sym['decision'])}
        <hr>
        <b>[非對稱]</b><br>
        Lower = {fmt2(asym['L'])} , Upper = {fmt2(asym['U'])} <br>
        Left = {fmt2(asym['half_L'])} , Right = {fmt2(asym['half_R'])} <br>
        Decision = {decision_html(asym['decision'])}
        </div>
        """,
            unsafe_allow_html=True,
        )

    # Tables
    results_df = build_results_df(res)
    work_df = build_workpapers_df(data)

    st.markdown("### 結果表 / Results")
    # use_container_width 改為 width="stretch"（新版語法）
    st.dataframe(results_df.astype(str), width="stretch")

    st.markdown("### 樣本明細 / Workpapers (with ej²)")
    # 防止 'Total' 轉換錯誤，統一轉為字串
    st.dataframe(work_df.astype(str), width="stretch")
    
    st.sidebar.markdown("---")
    teaching_mode = st.sidebar.checkbox("啟用教學模式 / Enable Teaching Mode", value=True)

    if teaching_mode:
        with st.expander("📘 教學模式 / Teaching Mode — 公式推導（對稱） / Symmetric formulas", expanded=True):
            i = res["inputs"]; n = res["n_used"]; sym = res["sym"]

            st.markdown("#### 變數意義對照表 / Variable Glossary")
            st.markdown("""
| Symbol | English | 中文 |
|:--:|:--|:--|
| *n* | Sample size | 樣本數 |
| *N* | Population size | 母體大小 |
| *ē* | Sample mean error | 樣本平均誤差 |
| *SD★* | Estimated population SD | 估計母體標準差 |
| *SD* | Sample SD (STDEV.S) | 樣本標準差 |
| *ZA* | ARIA Z-factor | 誤受險 Z 值 |
| *ZR* | ARIR Z-factor | 誤拒險 Z 值 |
| *TM* | Tolerable misstatement | 可容忍錯誤 |
| *E★* | Expected misstatement | 預期錯誤 |
| *CPI* | Confidence precision interval | 精確度區間（半寬） |
| *T̂* | Estimated total misstatement | 總誤差估計 |
| *L*, *U* | Lower / Upper bound | 區間下限 / 上限 |
            """)
            
            st.markdown("### 📐 完整公式與符號定義（Reference）")
            # [1] n (auto)
            st.markdown("#### [1] 樣本量 / Sample size (n)")
            st.markdown("**Formula**：  \n"
                        r"$$ n = \left[ \frac{SD^\star \cdot (Z_A + Z_R) \cdot N}{TM - E^\star} \right]^2 $$")
            st.markdown(f"**Substitution**：  \n"
                        f"= [ {fmt4(i['SD_star'])} × ({fmt4(i['ZA'])}+{fmt4(i['ZR'])}) × {i['N']:,} / ({fmt2(i['TM'])} − {fmt2(i['E_star'])}) ]²")
            st.markdown(f"**Result**：  \n"
                        f"= {res['n_auto_exact']:.10f} → ceil = **{res['n_auto_int']}**; used **n = {res['n_used']}**")

            # [2] T̂
            st.markdown("#### [2] 總誤差（點估計）/ Total misstatement (T̂)")
            st.markdown("**Formula**：  \n"
                        r"$$ \hat{T} = N \cdot \bar{e} $$")
            st.markdown(f"**Substitution**：  \n"
                        f"= {i['N']:,} × {fmt4(res['e_bar'])}")
            st.markdown(f"**Result**：  \n"
                        f"= **{fmt2(res['T_hat'])}**")

            # [3] CPI
            st.markdown("#### [3] 精確度區間（半寬）/ CPI (amount, halfwidth)")
            st.markdown("**Formula**（Excel-consistent; use sample SD）：  \n"
                        r"$$ \mathrm{CPI} = N \cdot Z_A \cdot SD \cdot \frac{1}{\sqrt{n}} \cdot \sqrt{\frac{N-n}{N}} $$")
            st.markdown(f"**Substitution**：  \n"
                        f"= {i['N']:,} × {fmt4(i['ZA'])} × {fmt4(res['SD_sample'])} × 1/√{n} × √(({i['N']:,}−{n})/{i['N']:,})")
            st.markdown(f"**Result**：  \n"
                        f"= **{fmt2(res['CPI_amt'])}**")

    # Export
    xlsx_bytes = export_to_excel(results_df, work_df)

    try:
        os.makedirs(os.path.dirname(FIXED_OUT_PATH), exist_ok=True)
        with open(FIXED_OUT_PATH, "wb") as f:
            f.write(xlsx_bytes)
        st.success(f"已另存到：{FIXED_OUT_PATH}")
    except Exception as e:
        st.warning(f"固定路徑另存失敗：{e}")

    st.download_button(
        label="下載 Excel 結果檔 / Download Results (.xlsx)",
        data=xlsx_bytes,
        file_name="output-CVS-DE.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    # ===== Teaching Mode (bottom, symmetric only) =====
    if teaching_mode:
        st.markdown("## 📘 教學模式 / Teaching Mode")

        tab1, tab2, tab3 = st.tabs([
            "👀 快速理解 (For students)",
            "📐 逐步公式 (In class)",
            "📚 完整定義 (Reference)"
        ])

        # ---------- Tab 1: 快速理解 ----------
        with tab1:
            st.markdown("""
    **一句話理解 CVS-DE：**

    > 我們用「樣本誤差」來推估「母體總誤差」，  
    > 並檢查在考慮審計風險後，是否仍低於可容忍錯誤（TM）。

    **你現在看到的結果重點：**
    - ✅ `T̂`：母體誤差的點估計
    - ✅ `CPI`：估計的不確定範圍
    - ✅ `U ≤ TM` → **接受（Accept）**
    - ✅ `U > TM` → **拒絕（Reject）**

    👉 不用背公式，先看結論。
            """)

        # ---------- Tab 2: 逐步公式 ----------
        with tab2:
            with st.expander("Step 1️⃣ 樣本量估計（Sample size）", expanded=False):
                st.latex(r"""
    n = \left[ \frac{SD^\star \cdot (Z_A + Z_R) \cdot N}{TM - E^\star} \right]^2
    """)
                st.caption("這一步決定要抽多少筆樣本。")

            with st.expander("Step 2️⃣ 點估計（Point estimate）"):
                st.latex(r"""
    \bar{e} = \frac{\sum e_i}{n}, \quad \hat{T} = N \cdot \bar{e}
    """)

        with st.expander("Step 3️⃣ 精確度區間（CPI）", expanded=False):

            st.markdown("**核心計算式（考試 / 判斷用）**")
            st.latex(r"""
        CPI = N \cdot Z_A \cdot \frac{SD}{\sqrt{n}} \sqrt{\frac{N-n}{N}}
        """
)
            # ✅ 次層 expander：只給想深究的人
            with st.expander("📐 推導與補充說明（進階）"):
                st.latex(r"""
        SD = \sqrt{\frac{\sum (e_i - \bar{e})^2}{n-1}}
        """)
                st.latex(r"""
        SE(\bar{e}) = \frac{SD^\star}{\sqrt{n}} \sqrt{\frac{N-n}{N}}
        """)
                st.markdown("""
        - FPC（finite population correction）只在 **抽樣比例不小** 時顯著  
        - 考試通常直接給 CPI，不要求推導
                """)

            with st.expander("Step 4️⃣ 決策（Decision rule）"):
                st.markdown("""
        - 若 **上限 U ≤ TM** → ✅ Accept  
        - 若 **上限 U > TM** → ❌ Reject  
                        """)

        # ---------- Tab 3: 完整定義 ----------
    with tab3:
        st.markdown("### 📚 完整公式與符號定義（Reference）")

        # ---------- 公式 ----------
        with st.expander("📐 完整公式（查閱用）", expanded=False):
            st.latex(r"n = \left[ \frac{SD^\star (Z_A + Z_R) N}{TM - E^\star} \right]^2")
            st.latex(r"\hat{T} = N \cdot \bar{e}")
            st.latex(r"CPI = N \cdot Z_A \cdot \frac{SD}{\sqrt{n}} \sqrt{\frac{N-n}{N}}")
            st.latex(r"L = \hat{T} - CPI,\quad U = \hat{T} + CPI")

        # ---------- Variable Glossary ----------
        with st.expander("🔤 Variable Glossary / 變數對照表", expanded=False):
            st.table(pd.DataFrame({
                "Symbol": ["n","N","ē","SD★","SD","ZA","ZR","TM","E★","CPI","T̂","L / U"],
                "Meaning (EN)": [
                    "Sample size",
                    "Population size",
                    "Sample mean error",
                    "Estimated population SD",
                    "Sample SD (STDEV.S)",
                    "ARIA Z-factor",
                    "ARIR Z-factor",
                    "Tolerable misstatement",
                    "Expected misstatement",
                    "Confidence precision interval (halfwidth)",
                    "Estimated total misstatement",
                    "Lower / Upper bound"
                ],
                "中文說明": [
                    "樣本數",
                    "母體大小",
                    "樣本平均誤差",
                    "估計母體標準差",
                    "樣本標準差",
                    "誤受險 Z 值",
                    "誤拒險 Z 值",
                    "可容忍錯誤",
                    "預期錯誤",
                    "精確度區間（半寬）",
                    "總誤差估計",
                    "區間下限／上限"
                ]
            }))

### Variable Glossary / 變數意義對照
st.markdown("""
| Symbol | Meaning (EN) | 中文說明 |
|:--:|:--|:--|
| *n* | Sample size | 樣本數 |
| *N* | Population size | 母體大小 |
| *ē* | Sample mean error | 樣本平均誤差 |
| *SD★* | Estimated population SD | 估計母體標準差 |
| *SD* | Sample SD (STDEV.S) | 樣本標準差 |
| *ZA* | ARIA Z-factor | 誤受險 Z 值 |
| *ZR* | ARIR Z-factor | 誤拒險 Z 值 |
| *TM* | Tolerable misstatement | 可容忍錯誤 |
| *E★* | Expected misstatement | 預期錯誤 |
| *CPI* | Confidence precision interval | 精確度區間（半寬） |
| *T̂* | Estimated total misstatement | 總誤差估計 |
| *L*, *U* | Lower / Upper bound | 區間下限 / 上限 |
""")

