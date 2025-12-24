# -*- coding: utf-8 -*-
"""
1111 New CVS-DE Testing-Streamlit.py
Streamlit app for Classical Variable Sampling (CVS-DE)
- Bilingual UI (中文/English)
- Upload Excel with sheet 'CVS-DE'
- Modes: Symmetric / Asymmetric / Both
- CPI uses sample SD (ddof=1), symmetric halfwidth = CPI (Excel-consistent)
- Export fixed path: H:\VS Code-Auditing\output-CVS-DE.xlsx + browser download
- Teaching Mode (bottom, collapsible): Variable Glossary + numbered formulas (Symmetric only)

Run:
    streamlit run "1111 New CVS-DE Testing-Streamlit.py"
"""

import io
import math
import os
from typing import Any, Dict, Optional, Tuple
import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

import datetime

# ===== Fixed output path (for local Excel export) =====
FIXED_OUT_PATH = r"H:\VS Code-Auditing\output-CVS-DE.xlsx"

# ===== Confidence level to (ZA, ZR) map (if you ever want auto-mode) =====
CONF_LEVEL_MAP: Dict[str, Tuple[float, float]] = {
    # key: "overall CL" → (ZA, ZR)  (not currently auto-used, but kept for extension)
    "99": (2.33, 2.58),
    "95": (1.65, 1.96),
    "90": (1.28, 1.64),
    "80": (0.84, 1.28),
}

# ===== Small formatting helpers =====
def fmt2(x: float) -> str:
    try:
        return f"{x:,.2f}"
    except Exception:
        return str(x)

def fmt4(x: float) -> str:
    try:
        return f"{x:,.4f}"
    except Exception:
        return str(x)

def fmt_int(x: float) -> str:
    try:
        return f"{int(round(x)):,}"
    except Exception:
        return str(x)

def safe_number(x: Any) -> Optional[float]:
    try:
        return float(x)
    except Exception:
        return None

# ===== Load sample block from uploaded Excel (sheet CVS-DE) =====
def load_sample_block_from_uploaded(file_buffer):
    raw = pd.read_excel(file_buffer, sheet_name="CVS-DE", header=None)

    def row_has_header(sr: pd.Series) -> bool:
        s = sr.astype(str).str.lower()
        return s.str.contains("account no", na=False).any() or s.str.contains("account number", na=False).any()

    header_idx_list = raw.index[raw.apply(row_has_header, axis=1)].tolist()
    if not header_idx_list:
        raise RuntimeError("找不到表頭列（需含 'Account NO.' 或 'Account Number'）")
    header_idx = header_idx_list[0]

    data = raw.iloc[header_idx + 1:, 0:4].copy()
    data.columns = [
        "Account Number",
        "Recorded Accounts",
        "Audited Accounts",
        "Factual Misstatement (ej)",
    ]

    # Trim spaces
    data["Account Number"] = data["Account Number"].astype(str).str.strip()

    # Stop at first row containing 'Total' in Account Number
    stop_idx = data.index[data["Account Number"].astype(str).str.contains("Total", case=False, na=False)]
    if len(stop_idx) > 0:
        data = data.loc[: stop_idx[0] - 1]

    # Cast numeric for main numeric columns
    for c in ["Recorded Accounts", "Audited Accounts", "Factual Misstatement (ej)"]:
        data[c] = pd.to_numeric(data[c], errors="coerce")

    # --- Stability patch: remove non-numeric Account Number rows (e.g., "Total", text, blanks) ---
    # This prevents Arrow / Streamlit from failing when converting to numeric types internally.
    data["AN_numeric"] = pd.to_numeric(data["Account Number"], errors="coerce")
    data = data[data["AN_numeric"].notna()].copy()
    data.drop(columns=["AN_numeric"], inplace=True)

    e_series = data["Factual Misstatement (ej)"].dropna()
    m = int(e_series.shape[0])
    sum_e = float(e_series.sum())
    visible_rows = int(data.shape[0])
    return data, e_series, m, sum_e, visible_rows

# ===== Core compute (Excel-consistent CPI & intervals) =====
def compute_all(
    SD_star: float, ZA: float, ZR: float, N: int, TM: float, E_star: float,
    e_series: pd.Series, m: int, sum_e: float, n_override: Optional[int]
) -> Dict[str, Any]:

    if TM <= E_star:
        raise ValueError("TM must be greater than E* to compute sample size (avoid division by zero).")

    # Auto sample size by formula
    n_auto_exact = ((SD_star * (ZA + ZR) * N) / (TM - E_star)) ** 2
    n_auto_int = int(math.ceil(n_auto_exact))

    if n_auto_int < 1:
        raise ValueError("Computed sample size n_auto is < 1, please check SD*, TM, E* settings.")

    # Use override if reasonable
    n = int(n_override) if (isinstance(n_override, int) and n_override >= 1) else n_auto_int
    if m > n:
        raise ValueError(f"Count of nonzero errors m={m} cannot exceed chosen sample size n={n}.")

    # Point estimate: ē = Σe / n (fill zero for non-error draws)
    e_bar = sum_e / n
    e_full = np.concatenate([e_series.values, np.zeros(max(0, n - m))])
    SD_sample = float(np.std(e_full, ddof=1)) if n > 1 else 0.0  # Excel STDEV.S

    # SE(mean) using SD* with FPC
    SE_mean = SD_star / math.sqrt(n) * math.sqrt((N - n) / N)

    # Total misstatement
    T_hat = N * e_bar

    # CPI（用樣本 SD，Excel 方法）
    CPI_amt = N * ZA * SD_sample / math.sqrt(n) * math.sqrt((N - n) / N)

    # Symmetric interval using CPI as half-width
    L_sym = T_hat - CPI_amt
    U_sym = T_hat + CPI_amt
    half_sym = CPI_amt

    # Asymmetric interval (lower = ZR, upper = ZA)
    L_asym = T_hat - N * ZR * SE_mean
    U_asym = T_hat + N * ZA * SE_mean
    half_L = T_hat - L_asym
    half_R = U_asym - T_hat

    # Decisions: compare upper bound vs TM
    dec_sym = "Accept" if U_sym <= TM else "Reject"
    dec_asym = "Accept" if U_asym <= TM else "Reject"

    return {
        "n_auto_exact": n_auto_exact,
        "n_auto_int": n_auto_int,
        "n_used": n,
        "inputs": {"SD_star": SD_star, "ZA": ZA, "ZR": ZR, "N": N, "TM": TM, "E_star": E_star, "m": m, "sum_e": sum_e},
        "e_bar": e_bar,
        "SD_sample": SD_sample,
        "SE_mean": SE_mean,
        "T_hat": T_hat,
        "CPI_amt": CPI_amt,
        "sym": {"Z_sym": None, "L": L_sym, "U": U_sym, "half": half_sym, "decision": dec_sym},
        "asym": {"ZA": ZA, "ZR": ZR, "L": L_asym, "U": U_asym, "half_L": half_L, "half_R": half_R, "decision": dec_asym},
    }

# ===== Results table (bilingual) =====
def build_results_table(res: Dict[str, Any]) -> pd.DataFrame:
    i = res["inputs"]
    sym = res["sym"]
    asym = res["asym"]

    rows = []

    # Block 1: Inputs + sample size
    rows.append(["n_auto", "Estimated sample size (auto)", "估計樣本量（公式）", fmt_int(res["n_auto_int"])])
    rows.append(["n_used", "Sample size actually used", "採用樣本量", fmt_int(res["n_used"])])
    rows.append(["N", "Population size", "母體大小", fmt_int(i["N"])])
    rows.append(["SD★", "Estimated population SD (SD★)", "估計母體標準差 SD★", fmt4(i["SD_star"])])
    rows.append(["ZA", "ARIA Z-factor", "誤受險 Z 值 ZA", fmt4(i["ZA"])])
    rows.append(["ZR", "ARIR Z-factor", "誤拒險 Z 值 ZR", fmt4(i["ZR"])])
    rows.append(["TM", "Tolerable misstatement (TM)", "可容忍錯誤 TM", fmt2(i["TM"])])
    rows.append(["E★", "Expected misstatement (E★)", "預期錯誤 E★", fmt2(i["E_star"])])

    # Block 2: Point estimate and SD
    rows.append(["ē", "Sample mean error", "樣本平均誤差 ē", fmt4(res["e_bar"])])
    rows.append(["SD", "Sample SD (STDEV.S)", "樣本標準差 SD", fmt4(res["SD_sample"])])
    rows.append(["SE(mean)", "Std. error of mean (with FPC)", "平均數標準誤（含有限母體修正）", fmt4(res["SE_mean"])])
    rows.append(["T̂", "Estimated total misstatement", "總誤差點估計 T̂ = N·ē", fmt2(res["T_hat"])])
    rows.append(["CPI", "Confidence precision interval (halfwidth)", "精確度區間（半寬 CPI）", fmt2(res["CPI_amt"])])

    # Block 3: Symmetric interval
    rows.append(["Sym_L", "Symmetric lower bound", "對稱信賴區間下限", fmt2(sym["L"])])
    rows.append(["Sym_U", "Symmetric upper bound", "對稱信賴區間上限", fmt2(sym["U"])])
    rows.append(["Sym_half", "Symmetric halfwidth (CPI)", "對稱半寬", fmt2(sym["half"])])
    rows.append(["Sym_decision", "Decision (Sym)", "決策（對稱）", sym["decision"]])

    # Block 4: Asymmetric interval
    rows.append(["Asym_L", "Asymmetric lower bound", "非對稱信賴區間下限", fmt2(asym["L"])])
    rows.append(["Asym_U", "Asymmetric upper bound", "非對稱信賴區間上限", fmt2(asym["U"])])
    rows.append(["Asym_half_L", "Asymmetric left halfwidth", "左側半寬", fmt2(asym["half_L"])])
    rows.append(["Asym_half_R", "Asymmetric right halfwidth", "右側半寬", fmt2(asym["half_R"])])
    rows.append(["Asym_decision", "Decision (Asym)", "決策（非對稱）", asym["decision"]])

    df = pd.DataFrame(rows, columns=["Key", "Item (EN)", "項目（中文）", "Value"])
    return df

# ===== Export to Excel (fixed path + download) =====
def export_to_excel(res_df: pd.DataFrame, work_df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "results"
    for r in dataframe_to_rows(res_df, index=False, header=True):
        ws1.append(r)

    ws2 = wb.create_sheet("workpapers")
    for r in dataframe_to_rows(work_df, index=False, header=True):
        ws2.append(r)

    # Apply accounting-like format to numeric-ish cells
    def set_thousand(ws):
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                try:
                    float(cell.value)
                    cell.number_format = '#,##0.00'
                except Exception:
                    pass

    set_thousand(ws1)
    set_thousand(ws2)

    # Save to bytes buffer
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    # Also save to fixed local path (for your H:\ usage)
    try:
        os.makedirs(os.path.dirname(FIXED_OUT_PATH), exist_ok=True)
        wb.save(FIXED_OUT_PATH)
    except Exception as e:
        # Silent fail is okay for competition; you still get download
        print(f"[WARN] Failed to save fixed path Excel: {e}")

    return bio.getvalue()

# ===== Streamlit App =====
def main():
    st.set_page_config(page_title="Classical Variable Sampling (CVS-DE)", layout="wide")

    st.title("📊 Classical Variable Sampling (CVS-DE)")
    st.markdown("### 審計抽樣互動平台 · Classical Variable Sampling for Auditing")

    # Top info / timestamp
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    st.caption(f"Run time / 執行時間：{now}")

    st.markdown(
        """
**說明 / Description**

本系統示範「傳統變量抽樣 (Classical Variable Sampling)」在審計抽樣中的應用，  
提供樣本量估計、點估計、精確度區間 (CPI)、對稱與非對稱信賴區間，以及決策判斷。

This app demonstrates Classical Variable Sampling for audit sampling,
including sample size estimation, point estimate, CPI, symmetric/asymmetric confidence intervals, and decisions.
"""
    )

    st.sidebar.header("Step 1. 上傳查核樣本 Excel / Upload Excel")
    uploaded_file = st.sidebar.file_uploader("請上傳含『CVS-DE』工作表之 Excel 檔案", type=["xlsx"])

    st.sidebar.header("Step 2. 輸入參數 / Input Parameters")

    colZA, colZR = st.sidebar.columns(2)
    ZA = colZA.number_input("ZA (誤受險因子)", value=1.28, step=0.01, format="%.2f")
    ZR = colZR.number_input("ZR (誤拒險因子)", value=1.15, step=0.01, format="%.2f")

    colN, colTM = st.sidebar.columns(2)
    N = int(colN.number_input("母體大小 N", value=4000, step=1))
    TM = colTM.number_input("可容忍錯誤 TM", value=21000.0, step=100.0)

    colE, colSD = st.sidebar.columns(2)
    E_star = colE.number_input("預期錯誤 E★", value=1500.0, step=100.0)
    SD_star = colSD.number_input("估計母體標準差 SD★", value=20.0, step=0.1)

    st.sidebar.markdown("---")
    st.sidebar.write("**樣本量控管 / Sample size control**")
    n_override = st.sidebar.number_input("確認樣本量 n（留 0 代表採用公式預設）", min_value=0, step=1, value=100)

    st.sidebar.markdown("---")
    mode = st.sidebar.radio(
        "信賴區間模式 / Interval mode",
        options=["Symmetric only", "Asymmetric only", "Both (並列比較)"],
        index=2,
    )

    st.sidebar.markdown("---")
    teaching_mode = st.sidebar.checkbox("啟用教學模式 / Enable Teaching Mode", value=True)

    st.markdown("## 🔍 分析結果 / Analysis Results")

    if uploaded_file is None:
        st.info("請從左側上傳含 CVS-DE 工作表的 Excel 檔。 / Please upload an Excel file with sheet 'CVS-DE'.")
        return

    try:
        work_df, e_series, m, sum_e, visible_rows = load_sample_block_from_uploaded(uploaded_file)
    except Exception as e:
        st.error(f"讀取樣本資料時發生錯誤 / Error while reading sample block: {e}")
        return

    # Compute
    n_override_int = int(n_override) if n_override > 0 else None

    try:
        res = compute_all(SD_star, ZA, ZR, N, TM, E_star, e_series, m, sum_e, n_override_int)
    except Exception as e:
        st.error(f"計算過程發生錯誤 / Error in computation: {e}")
        return

    # Results table and workpapers
    results_df = build_results_table(res)
    left, right = st.columns(2)

    with left:
        st.markdown("### 📑 結果摘要 / Summary table")
        st.dataframe(results_df.astype(str), width="stretch")

    with right:
        st.markdown("### 📂 查核樣本 / Audit sample (workpapers)")
        st.caption(f"可見列數 (visible rows) = {visible_rows}, 非零誤差筆數 m = {m}")
        st.dataframe(work_df.astype(str), width="stretch")

    # Interval display block (according to mode)
    sym = res["sym"]
    asym = res["asym"]

    if mode in ["Symmetric only", "Both (並列比較)"]:
        st.markdown("#### 🎯 對稱信賴區間 / Symmetric interval")
        st.write(f"下限 L = {fmt2(sym['L'])}, 上限 U = {fmt2(sym['U'])}")
        st.write(f"半寬 (CPI) = {fmt2(sym['half'])}")
        st.write(f"決策 / Decision: **{sym['decision']}**")

    if mode in ["Asymmetric only", "Both (並列比較)"]:
        st.markdown("#### ⚖️ 非對稱信賴區間 / Asymmetric interval")
        st.write(f"下限 L = {fmt2(asym['L'])}, 上限 U = {fmt2(asym['U'])}")
        st.write(f"左側半寬 = {fmt2(asym['half_L'])}, 右側半寬 = {fmt2(asym['half_R'])}")
        st.write(f"決策 / Decision: **{asym['decision']}**")

    # Export section
    st.markdown("---")
    st.markdown("### 💾 匯出結果 / Export to Excel")

    excel_bytes = export_to_excel(results_df, work_df)
    st.download_button(
        label="⬇ 下載結果 Excel / Download results Excel",
        data=excel_bytes,
        file_name="output-CVS-DE.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    st.caption(f"本機另存路徑（若成功）：{FIXED_OUT_PATH}")

    # Teaching mode
    if teaching_mode:
        st.markdown("## 📘 教學模式 / Teaching Mode")

        tab1, tab2, tab3 = st.tabs([
            "👀 快速理解 (For students)",
            "📐 逐步公式 (In class)",
            "📚 完整定義 (Reference)"
        ])

        # ---------- Tab 1: 快速理解 ----------
        with tab1:
            st.markdown("""
    **一句話理解 CVS-DE：**

    > 我們用「樣本誤差」來推估「母體總誤差」，  
    > 並檢查在考慮審計風險後，是否仍低於可容忍錯誤（TM）。

    **你現在看到的結果重點：**
    - ✅ `T̂`：母體誤差的點估計
    - ✅ `CPI`：估計的不確定範圍
    - ✅ `U ≤ TM` → **接受（Accept）**
    - ✅ `U > TM` → **拒絕（Reject）**

    👉 不用背公式，先看結論。
            """)

        # ---------- Tab 2: 逐步公式 ----------
        with tab2:
            with st.expander("Step 1️⃣ 樣本量估計（Sample size）", expanded=False):
                st.latex(r"""
    n = \left[ \frac{SD^\star \cdot (Z_A + Z_R) \cdot N}{TM - E^\star} \right]^2
    """)
                st.caption("這一步決定要抽多少筆樣本。")

            with st.expander("Step 2️⃣ 點估計（Point estimate）"):
                st.latex(r"""
    \bar{e} = \frac{\sum e_i}{n}, \quad \hat{T} = N \cdot \bar{e}
    """)

        with st.expander("Step 3️⃣ 精確度區間（CPI）", expanded=False):

            st.markdown("**核心計算式（考試 / 判斷用）**")
            st.latex(r"""
        CPI = N \cdot Z_A \cdot \frac{SD}{\sqrt{n}} \sqrt{\frac{N-n}{N}}
        """
)
            # ✅ 次層 expander：只給想深究的人
            with st.expander("📐 推導與補充說明（進階）"):
                st.latex(r"""
        SD = \sqrt{\frac{\sum (e_i - \bar{e})^2}{n-1}}
        """)
                st.latex(r"""
        SE(\bar{e}) = \frac{SD^\star}{\sqrt{n}} \sqrt{\frac{N-n}{N}}
        """)
                st.markdown("""
        - FPC（finite population correction）只在 **抽樣比例不小** 時顯著  
        - 考試通常直接給 CPI，不要求推導
                """)

            with st.expander("Step 4️⃣ 決策（Decision rule）"):
                st.markdown("""
        - 若 **上限 U ≤ TM** → ✅ Accept  
        - 若 **上限 U > TM** → ❌ Reject  
                        """)

        # ---------- Tab 3: 完整定義 ----------
    with tab3:
        st.markdown("### 📚 完整公式與符號定義（Reference）")

        # ---------- 公式 ----------
        with st.expander("📐 完整公式（查閱用）", expanded=False):
            st.latex(r"n = \left[ \frac{SD^\star (Z_A + Z_R) N}{TM - E^\star} \right]^2")
            st.latex(r"\hat{T} = N \cdot \bar{e}")
            st.latex(r"CPI = N \cdot Z_A \cdot \frac{SD}{\sqrt{n}} \sqrt{\frac{N-n}{N}}")
            st.latex(r"L = \hat{T} - CPI,\quad U = \hat{T} + CPI")

        # ---------- Variable Glossary ----------
        with st.expander("🔤 Variable Glossary / 變數對照表", expanded=False):
            st.table(pd.DataFrame({
                "Symbol": ["n","N","ē","SD★","SD","ZA","ZR","TM","E★","CPI","T̂","L / U"],
                "Meaning (EN)": [
                    "Sample size",
                    "Population size",
                    "Sample mean error",
                    "Estimated population SD",
                    "Sample SD (STDEV.S)",
                    "ARIA Z-factor",
                    "ARIR Z-factor",
                    "Tolerable misstatement",
                    "Expected misstatement",
                    "Confidence precision interval (halfwidth)",
                    "Estimated total misstatement",
                    "Lower / Upper bound"
                ],
                "中文說明": [
                    "樣本數",
                    "母體大小",
                    "樣本平均誤差",
                    "估計母體標準差",
                    "樣本標準差",
                    "誤受險 Z 值",
                    "誤拒險 Z 值",
                    "可容忍錯誤",
                    "預期錯誤",
                    "精確度區間（半寬）",
                    "總誤差估計",
                    "區間下限／上限"
                ]
            }))


### Variable Glossary / 變數意義對照

| Symbol | Meaning (EN) | 中文說明 |
|:--:|:--|:--|
| *n* | Sample size | 樣本數 |
| *N* | Population size | 母體大小 |
| *ē* | Sample mean error | 樣本平均誤差 |
| *SD★* | Estimated population SD | 估計母體標準差 |
| *SD* | Sample SD (STDEV.S) | 樣本標準差 |
| *ZA* | ARIA Z-factor | 誤受險 Z 值 |
| *ZR* | ARIR Z-factor | 誤拒險 Z 值 |
| *TM* | Tolerable misstatement | 可容忍錯誤 |
| *E★* | Expected misstatement | 預期錯誤 |
| *CPI* | Confidence precision interval | 精確度區間（半寬） |
| *T̂* | Estimated total misstatement | 總誤差估計 |
| *L*, *U* | Lower / Upper bound | 區間下限 / 上限 |
